'''
Created on 2018年1月15日

@author: benlei
'''
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from .models import ForeignDataTestCase
from openpyxl.compat import range
from statsmodels.discrete.tests.test_sandwich_cov import filepath
            
def importExcel(user,filepath):
    wb = load_workbook(filepath)
    ws = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(ws[0])
    headers = ['TRAN_CODE', 'REMARK1', 'REMARK2', 'REMARK3', 'REMARK4', 'BODYXML']
    lists = []
    for row in range(4, 5):
        r = {}
        for col in range(1, len(headers) + 1):
            key = headers[col - 1]
            r[key] = ws.cell(row=row, column=col).value
        lists.append(r)
    sqllist = []
    for cell in lists:
        # for header in headers:
        TRAN_CODE = cell['TRAN_CODE']
        REMARK1 = cell['REMARK1']
        REMARK2 = cell['REMARK2']
        REMARK3 = cell['REMARK3']
        REMARK4 = cell['REMARK4']
        BODYXML = cell['BODYXML']
        sql = ForeignDataTestCase(TRAN_CODE =TRAN_CODE,
                                    REMARK1 = REMARK1,
                                    REMARK2 = REMARK2,
                                    REMARK3 = REMARK3,
                                    REMARK4 = REMARK4,
                                    BODYXML = BODYXML,
                                    editor=user)
        sqllist.append(sql)
    ForeignDataTestCase.objects.bulk_create(sqllist)
    